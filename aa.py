import openpyxl
import pymysql
import GetFile as g
#資料庫連線設定
db = pymysql.connect(host='localhost', port=3306, user='root', passwd='123456', db='sys')
#建立操作游標
cursor = db.cursor()
#打開excel檔並以wb儲存
wb = openpyxl.load_workbook('goodFinal.xlsx')
sheet = wb.worksheets[0]
#commit至資料庫的方法
def commit():
    try:
      #提交修改
      db.commit()
      print('success')
    except:
      #發生錯誤時停止執行SQL
      db.rollback()
      print('error')
#開始跑資料庫
for i in range(1, sheet.max_row+1):
    print("row="+str(i))
    for j in range(1, sheet.max_column+1):
        ing = str(sheet.cell(row=i, column=j).value) #ing是每一格的值
        id = 0
        ingre_name, what_it_does, irritancy, comedogenicity, rank, alias = g.incide2(ing) #取得各元素
        if (rank != 'nothing'):
            sql = "SELECT id FROM ingres_alia where name =(%s)"
            val = ingre_name
            cursor.execute(sql, val)
            id = cursor.fetchone()
            if(id == None):
                sql = "SELECT id FROM ingres where name = (%s)"
                val = ingre_name
                cursor.execute(sql, val)
                id = cursor.fetchone()
                if(id == None):
                    sql = "INSERT INTO ingres (name,irr,com,tier) VALUES (%s, %s, %s, %s)"
                    val = (ingre_name, irritancy, comedogenicity, rank)
                    cursor.execute(sql, val)
                    commit()
                else:
                    print(ingre_name, "duplicate ingre")
                    continue
                sql = "SELECT id FROM ingres where name = (%s)"
                val = ingre_name
                cursor.execute(sql, val)
                id = cursor.fetchone()[0]
                if (len(alias) != 0):
                    for a in alias:
                        temp = a
                        sql = "INSERT INTO ingres_alia (name,id) VALUES (%s, %s)"
                        val = (a, id)
                        cursor.execute(sql, val)
                        commit()
                if (len(what_it_does) != 0):
                    for a in what_it_does:
                        sql = "SELECT id FROM whattheydo where func = (%s)"
                        val = (a)
                        cursor.execute(sql, val)
                        func_id = cursor.fetchone()
                        if(func_id == None):
                            sql = "INSERT INTO whattheydo (func) VALUES (%s)"
                            val = (a)
                            cursor.execute(sql, val)
                            commit()
                    for a in what_it_does:
                        sql = "SELECT id FROM whattheydo where func = (%s)"
                        val = (a)
                        cursor.execute(sql, val)
                        func_id = cursor.fetchone()[0]
                        sql = "INSERT INTO funcs (ingres_id,func_id) VALUES (%s, %s)"
                        val = (id, func_id)
                        cursor.execute(sql, val)
                        commit()
            else:
                print(ingre_name, "duplicate alia")
                continue
        else:
            print(ingre_name, "not found")
            continue
