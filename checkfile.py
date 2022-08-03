import openpyxl, requests, bs4, time

wb = openpyxl.load_workbook('goodFinal.xlsx')
sheet = wb.worksheets[0]
count = 0
for i in range(1, sheet.max_row+1):
    print("row="+str(i))
    for j in range(1, sheet.max_column+1):
        ing = str(sheet.cell(row=i, column=j).value)
        search = ''
        ing = ing.replace('\u200b', "")
        for a in ing:
            if (a.isalnum()):
                search += a.lower()
            elif (a == "/" or a == "-" or a == "," or a == " " or a == "[" or a == "]"):
                search += "-"
        re = requests.get('https://incidecoder.com/ingredients/' + search)
        soup = bs4.BeautifulSoup(re.text, "html.parser")
        try:
            name = soup.find("div", class_="ourtake grey1")
            if (name.text == 'icky'):
                print(ing + " is icky.")
                count+=1
        except:
            try:
                soup_list = soup.find_all("div", class_="itemprop")
            except:
                print(ing + ":[], '', ''")
            try:
                what_it_does = []
                irritancy, comedogenicity = '', ''
                for x in range(len(soup_list)):
                    try:
                        what_it_do = soup_list[x].select("a")
                        for y in range(len(what_it_do)):
                            what_it_does.append(what_it_do[y].getText())
                    except:
                        print('help')
                    if (soup_list[x].select("span") == 'Irritancy: '):
                        irritancies = soup_list[x].select("span")
                        for y in range(len(irritancies)):
                            if (y == 1):
                                irritancy = irritancies[y].getText()
                    elif (soup_list[x].select("span") == 'Comedogenicity: '):
                        comedogenicities = soup_list[x].select("span")
                        for y in range(len(comedogenicities)):
                            if (y == 1):
                                comedogenicity = comedogenicities[j].getText()
                if ('3' in irritancy or '3' in comedogenicity):
                    print(ing + what_it_does + "," + irritancy + "," + comedogenicity)
                    count+=1
            except:
                print(ing + ":[], '', ''")
print(count)
