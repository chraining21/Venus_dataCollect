import openpyxl,requests,bs4

wb = openpyxl.load_workbook('goodFinal.xlsx')
sheet = wb.worksheets[0]
for i in range(1, sheet.max_row):
    print("row="+str(i))
    for j in range(1, 7):
        ing = str(sheet.cell(row=i, column=j).value)
        search = ''
        count = 0
        ing = ing.replace('\u200b', "")
        for a in ing:
            if(a.isalnum()):
                count=0
                search+=a.lower()
            elif(a=="/" or a=="-" or a=="," or a==" " or a=="[" or a=="]"):
                search += "-"
        re = requests.get('https://incidecoder.com/ingredients/' + search)
        soup = bs4.BeautifulSoup(re.text, "html.parser")
        aloha = soup.find("h1")
        if (aloha.text == "Sorry, this page does not seem to exist"):
            print("["+str(i)+","+str(j)+"]"+search)