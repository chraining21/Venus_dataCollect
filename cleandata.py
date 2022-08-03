import openpyxl
n_wb = openpyxl.Workbook();
n_sh = n_wb.active
n_sh.title = "complete_Ingrdient"
_ings = []
row = 1
col = 1
wb = openpyxl.load_workbook('Good.xlsx')
sheet = wb.worksheets[0]
for i in range(1, sheet.max_row+1):
    for j in range(1, sheet.max_column+1):
        ing = str(sheet.cell(row = i, column = j).value)
        ing = ing.lower()
        if(not (ing in _ings)):
            _ings.append(ing)
for i in _ings:
    if (col == 7):
        row += 1
        col = 1
    n_sh.cell(row, col).value = i
    col += 1
n_wb.save("goodFinal.xlsx")